#! python3

import csv
from openpyxl import load_workbook, Workbook
from openpyxl.drawing import Image

currentEvent = None
currentCategory = None

with open('masters.csv', newline='') as csvfile:
    reader = csv.DictReader(csvfile)
    for row in reader:
        if currentEvent != row['Event']:
            if currentEvent is not None:
                #  Finish last event
                # remove extra sheets
                for sheet in sheets[wsCount:]:
                    output.remove_sheet(sheet)

                # save workbook
                output.save('output/{}.xlsx'.format(currentEvent))

            
            # Start new event
            currentEvent = row['Event']
            print(currentEvent)
            output = load_workbook('masters/{}.xlsx'.format(row['Event'].lower()))
            sheets = [sheet for sheet in output]
            wsCount = 0
        
        if currentCategory != row['Category']:
            # Start new spreadsheet
            currentCategory = row['Category']
            ws = sheets[wsCount]
            wsCount += 1

            # Add image
            img = Image('nbta.png')
            # place image relative to top left corner of spreadsheet
            img.drawing.top = 10
            img.drawing.left = 30
            ws.add_image(img)

            ws.title = row['Category'][:26]
            ws['C3'] = row['Category']

            i = 5

        # Add names to worksheet
        ws['B{}'.format(i)] = row['Name']
        i += 1

#  Finish last event
# remove extra sheets
for sheet in sheets[wsCount:]:
    output.remove_sheet(sheet)

# save workbook
output.save('output/{}.xlsx'.format(currentEvent))
